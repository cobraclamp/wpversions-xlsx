#!/usr/local/bin/python2.7
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font
)
import versionstatus

class Xlsx():

    XLSX_TITLE = "WordPress Versions"

    COLS = [
        "Domain",
        "Version",
        "Version Date",
        "Version Separation (Days)",
        "Days Since Update",
        "Severity",
    ]

    DICT_MAP = {
        "Domain": "domain",
        "Version": "version",
        "Version Date": "version_date",
        "Version Separation (Days)": "version_dist",
        "Days Since Update": "days_since",
        "Severity": "severity"
    }

    STYLES = {
        "SEVERE": {
            "fill": PatternFill(
                fill_type="solid",
                start_color="EF3A3A"
            ),
        },
        "GOOD": {
            "fill": PatternFill(
                fill_type="solid",
                start_color="92D964"
            )
        },
        "MEDIUM": {
            "fill": PatternFill(
                fill_type="solid",
                start_color="F3D11D"
            )
        }
    }

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid",start_color="DDDDDD")

    def __init__(self, wpversions, dest="wpversions"):
        now = datetime.now().strftime("%d_%m_%Y")
        self.wpversions = wpversions
        self.dest = dest + "_" + now + ".xlsx"
        self.wb = Workbook()

    def write_excel(self):
        ws = self.wb.active
        ws.title = self.XLSX_TITLE

        row_count = 1
        for row, domain in enumerate(self.wpversions):
            for i, col in enumerate(self.COLS):
                if row == 0:
                    ws.cell(row=1, column=i+1).value = self.COLS[i]
                    ws.cell(row=1, column=i+1).font = self.header_font
                    ws.cell(row=1, column=i+1).fill = self.header_fill
                ws.cell(row=(row+2), column=(i+1)).value = domain[self.DICT_MAP[self.COLS[i]]]
                if self.STYLES[domain["severity"]]:
                    ws.cell(row=(row+2), column=(i+1)).fill = self.STYLES[domain["severity"]]["fill"]
            row_count += 1

        sort_column = "D2:D%d" % (row_count)
        ws.auto_filter.add_sort_condition(sort_column)

        self.wb.save(filename=self.dest)
